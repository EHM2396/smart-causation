"""
Router: /tipos-comprobante – CRUD + carga masiva de tipos de comprobante.
Requiere autenticación; filtra por empresa activa.
"""

from __future__ import annotations

from io import BytesIO
from typing import Annotated

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from api.dependencies import get_empresa_activa
from api.schemas import TipoComprobanteCreate, TipoComprobanteOut
from db.models.auth import Empresa
from db.session import get_db
from services import tipos_service

router = APIRouter(prefix="/tipos-comprobante", tags=["Tipos de Comprobante"])

DB = Annotated[Session, Depends(get_db)]
EmpresaActiva = Annotated[Empresa, Depends(get_empresa_activa)]


@router.get("/", response_model=list[TipoComprobanteOut])
def get_tipos(db: DB, empresa: EmpresaActiva):
    return tipos_service.listar_todos(db, empresa_id=empresa.id)


@router.get("/opciones", response_model=list[dict])
def get_opciones(db: DB, empresa: EmpresaActiva):
    return tipos_service.listar_como_opciones(db, empresa_id=empresa.id)


@router.get("/plantilla")
def descargar_plantilla(empresa: EmpresaActiva):
    """Descarga un Excel de ejemplo con el formato esperado para importación."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tipos Comprobante"

    headers = ["codigo", "titulo"]
    ejemplos = [
        ["1",  "Comprobante de egreso"],
        ["2",  "Recibo de caja"],
        ["3",  "Nota de contabilidad"],
        ["4",  "Nota de ajuste"],
        ["12", "Ajustes contables"],
        ["13", "Compras nacionales"],
        ["14", "Cuentas por cobrar"],
        ["22", "Nómina"],
        ["FEC","Factura electrónica de compra"],
    ]

    header_fill = PatternFill("solid", fgColor="1C6B4A")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_data in ejemplos:
        ws.append(row_data)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 36

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_tipos_comprobante.xlsx"},
    )


def _detectar_fila_header_tc(raw: pd.DataFrame, max_scan: int = 12) -> int | None:
    """
    Detecta la fila de encabezados en archivos exportados de SIIGO (tipos comprobante).
    Busca la fila que contiene "Código del comprobante" en la primera celda (con posible
    corrupción de encoding → la celda contiene "digo").
    """
    for i in range(min(max_scan, len(raw))):
        first = str(raw.iloc[i, 0]).strip().lower()
        if "digo" in first and len(first) > 5:
            return i
    return None


@router.post("/cargar-excel")
async def cargar_excel(
    db: DB,
    empresa: EmpresaActiva,
    archivo: UploadFile = File(...),
):
    """
    Importa tipos de comprobante desde Excel. Acepta dos formatos:
      - Exportación directa de SIIGO (con filas de metadata al inicio).
      - Plantilla propia (columnas: codigo, titulo).
    """
    contenido = await archivo.read()
    try:
        raw = pd.read_excel(BytesIO(contenido), dtype=str, header=None)
    except Exception as e:
        raise HTTPException(400, f"No se pudo leer el archivo: {e}")

    # SIIGO siempre tiene filas de metadata antes del encabezado (header_row > 0).
    header_row = _detectar_fila_header_tc(raw)
    es_siigo = header_row is not None and header_row > 0

    if es_siigo:
        # Formato SIIGO: Código del comprobante(0), Título comprobante(1), Editar(2-ignorar)
        df = raw.iloc[header_row + 1:].reset_index(drop=True)
        COL_CODIGO = 0
        COL_TITULO = 1
    else:
        try:
            df = pd.read_excel(BytesIO(contenido), dtype=str)
        except Exception as e:
            raise HTTPException(400, f"No se pudo leer el archivo: {e}")
        df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
        for col in ("codigo", "titulo"):
            if col not in df.columns:
                raise HTTPException(400, f"El archivo debe tener una columna '{col}'.")

    insertados = 0
    actualizados = 0
    omitidos = 0
    errores: list[dict] = []

    for idx, row in df.iterrows():
        fila = int(idx) + (header_row + 2 if es_siigo else 2)

        if es_siigo:
            codigo = str(row.iloc[COL_CODIGO] if COL_CODIGO < len(row) else "").strip()
            titulo = str(row.iloc[COL_TITULO] if COL_TITULO < len(row) else "").strip()
        else:
            codigo = str(row.get("codigo", "")).strip()
            titulo = str(row.get("titulo", "")).strip()

        if not codigo or codigo == "nan":
            continue
        if not titulo or titulo == "nan":
            continue

        # Ignorar filas de pie de página SIIGO (ej: "Procesado en: Junio 27 2026")
        if codigo.lower().startswith("procesado"):
            omitidos += 1
            continue

        try:
            _, creado = tipos_service.upsert_tipo(
                db, codigo=codigo, titulo=titulo, empresa_id=empresa.id
            )
            if creado:
                insertados += 1
            else:
                actualizados += 1
        except Exception as e:
            errores.append({"fila": fila, "error": str(e)})

    db.commit()
    return {
        "insertados": insertados,
        "actualizados": actualizados,
        "omitidos": omitidos,
        "formato": "siigo" if es_siigo else "plantilla",
        "errores": errores,
    }


@router.post("/", response_model=TipoComprobanteOut, status_code=201)
def post_tipo(body: TipoComprobanteCreate, db: DB, empresa: EmpresaActiva):
    if tipos_service.buscar_por_codigo(db, body.codigo, empresa_id=empresa.id):
        raise HTTPException(409, f"El código {body.codigo!r} ya existe")
    tipo = tipos_service.crear_tipo(db, codigo=body.codigo, titulo=body.titulo, empresa_id=empresa.id)
    db.commit()
    return tipo
