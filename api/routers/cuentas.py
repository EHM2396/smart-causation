"""
Router: /cuentas – CRUD + carga masiva del plan de cuentas PUC.
Requiere autenticación; filtra por empresa activa.
"""

from __future__ import annotations

from io import BytesIO
from typing import Annotated

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from api.dependencies import get_empresa_activa
from api.schemas import CuentaCreate, CuentaOpcion, CuentaOut
from db.models.auth import Empresa
from db.session import get_db
from services import cuentas_service

router = APIRouter(prefix="/cuentas", tags=["Cuentas Contables"])

DB = Annotated[Session, Depends(get_db)]
EmpresaActiva = Annotated[Empresa, Depends(get_empresa_activa)]


@router.get("/gasto", response_model=list[CuentaOpcion])
def get_cuentas_gasto(db: DB, empresa: EmpresaActiva):
    return cuentas_service.listar_cuentas_gasto(db, empresa_id=empresa.id)


@router.get("/pago", response_model=list[CuentaOpcion])
def get_metodos_pago(db: DB, empresa: EmpresaActiva):
    return cuentas_service.listar_metodos_pago(db, empresa_id=empresa.id)


@router.get("/sugerencias", response_model=list[dict])
def get_sugerencias(
    descripcion: Annotated[str, Query(min_length=3)],
    db: DB,
    empresa: EmpresaActiva,
    max: int = 3,
):
    return cuentas_service.buscar_cuentas_sugeridas(db, descripcion, max, empresa_id=empresa.id)


@router.get("/plantilla")
def descargar_plantilla(empresa: EmpresaActiva):
    """Descarga un Excel de ejemplo con el formato esperado para importación."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan de Cuentas"

    headers = ["codigo", "nombre", "fiscal"]
    ejemplos = [
        ["51050505", "Honorarios servicios profesionales",      "NO"],
        ["51050510", "Asesorías jurídicas",                     "NO"],
        ["51100505", "Arrendamiento oficinas",                  "NO"],
        ["51200505", "Publicidad y propaganda",                 "NO"],
        ["23650500", "Retención en la fuente por pagar",        "NO"],
        ["24080500", "IVA por pagar",                           "NO"],
        ["22050505", "Proveedores nacionales",                  "NO"],
        ["11100501", "Caja general",                            "NO"],
        ["11200501", "Banco cuenta corriente",                  "NO"],
    ]

    header_fill = PatternFill("solid", fgColor="1C6B4A")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_data in ejemplos:
        ws.append(row_data)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 10

    ws.append([])
    nota = ws.cell(row=len(ejemplos) + 3, column=1, value="fiscal: SI o NO (default NO). Usar 8 dígitos para cuentas auxiliares.")
    nota.font = Font(italic=True, color="666666")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_plan_cuentas.xlsx"},
    )


def _normalizar_col(col: str) -> str:
    import unicodedata
    col = col.strip().lower().replace(" ", "_")
    col = unicodedata.normalize("NFD", col)
    col = "".join(c for c in col if unicodedata.category(c) != "Mn")
    return col


def _detectar_fila_header(raw: pd.DataFrame, max_scan: int = 12) -> int | None:
    """
    Detecta la fila de encabezados en archivos exportados de SIIGO.
    SIIGO coloca metadata de empresa arriba; el encabezado real contiene
    una celda que termina en 'digo' (de 'Código', con posible corrupción de encoding).
    Retorna el índice de fila o None si no se detecta el formato SIIGO.
    """
    for i in range(min(max_scan, len(raw))):
        first = str(raw.iloc[i, 0]).strip().lower()
        if first.endswith("digo") or first.startswith("c") and "digo" in first:
            return i
    return None


@router.post("/cargar-excel")
async def cargar_excel(
    db: DB,
    empresa: EmpresaActiva,
    archivo: UploadFile = File(...),
):
    """
    Importa cuentas PUC desde Excel. Acepta dos formatos:
      - Exportación directa de SIIGO (con filas de metadata al inicio).
      - Plantilla propia (columnas: codigo, nombre, fiscal).
    Filtros obligatorios: código de 8 dígitos + Nivel agrupación = Transaccional.
    """
    contenido = await archivo.read()
    try:
        raw = pd.read_excel(BytesIO(contenido), dtype=str, header=None)
    except Exception as e:
        raise HTTPException(400, f"No se pudo leer el archivo: {e}")

    # ── Detectar formato ───────────────────────────────────────────────────
    # SIIGO siempre tiene filas de metadata antes del encabezado (header_row > 0).
    # La plantilla propia tiene encabezados en fila 0 (header_row == 0 → no es SIIGO).
    header_row = _detectar_fila_header(raw)
    es_siigo = header_row is not None and header_row > 0

    if es_siigo:
        # Formato SIIGO: columnas por posición (el encoding corrompe los nombres)
        # Orden: Código(0), Nombre(1), Categoría(2), Clase(3), Relación con(4),
        #        Maneja vencimientos(5), Diferencia fiscal(6), Activo(7), Nivel agrupación(8)
        df = raw.iloc[header_row + 1:].reset_index(drop=True)
        COL_CODIGO  = 0
        COL_NOMBRE  = 1
        COL_FISCAL  = 6   # Diferencia fiscal
        COL_NIVEL   = 8   # Nivel agrupación
        tiene_nivel = raw.shape[1] > COL_NIVEL
    else:
        # Formato plantilla: leer normalmente con encabezados
        try:
            df = pd.read_excel(BytesIO(contenido), dtype=str)
        except Exception as e:
            raise HTTPException(400, f"No se pudo leer el archivo: {e}")
        df.columns = [_normalizar_col(c) for c in df.columns]
        if "codigo" not in df.columns:
            raise HTTPException(400, "El archivo debe tener una columna 'codigo'.")
        if "nombre" not in df.columns:
            raise HTTPException(400, "El archivo debe tener una columna 'nombre'.")
        tiene_nivel = "nivel_agrupacion" in df.columns

    insertados = 0
    actualizados = 0
    omitidos_codigo = 0
    omitidos_nivel  = 0
    errores: list[dict] = []

    for idx, row in df.iterrows():
        fila = int(idx) + (header_row + 2 if es_siigo else 2)

        if es_siigo:
            codigo = str(row.iloc[COL_CODIGO] if COL_CODIGO < len(row) else "").strip()
            nombre = str(row.iloc[COL_NOMBRE] if COL_NOMBRE < len(row) else "").strip()
            fiscal_raw = str(row.iloc[COL_FISCAL] if COL_FISCAL < len(row) else "").strip().lower()
            nivel_raw  = str(row.iloc[COL_NIVEL]  if tiene_nivel and COL_NIVEL < len(row) else "").strip().lower()
            fiscal = fiscal_raw in ("sí", "si", "s", "1", "true", "yes")
        else:
            codigo = str(row.get("codigo", "")).strip()
            nombre = str(row.get("nombre", "")).strip()
            fiscal_raw = str(row.get("fiscal", "NO")).strip().upper()
            fiscal = fiscal_raw in ("SI", "SÍ", "S", "1", "TRUE", "YES")
            nivel_raw = str(row.get("nivel_agrupacion", "")).strip().lower() if tiene_nivel else ""

        if not codigo or codigo == "nan" or not nombre or nombre == "nan":
            continue

        # ── Filtro 1: código exactamente 8 dígitos ────────────────────────
        if not (codigo.isdigit() and len(codigo) == 8):
            omitidos_codigo += 1
            continue

        # ── Filtro 2: nivel agrupación = Transaccional ────────────────────
        if tiene_nivel and nivel_raw and nivel_raw != "transaccional":
            omitidos_nivel += 1
            continue

        try:
            _, creado = cuentas_service.upsert_cuenta(
                db, codigo=codigo, nombre=nombre, fiscal=fiscal, empresa_id=empresa.id,
            )
            if creado:
                insertados += 1
            else:
                actualizados += 1
        except Exception as e:
            db.rollback()  # recuperar sesión tras flush fallido
            errores.append({"fila": fila, "error": str(e)})

    if insertados + actualizados > 0:
        db.commit()
    return {
        "insertados": insertados,
        "actualizados": actualizados,
        "omitidos_codigo": omitidos_codigo,
        "omitidos_nivel": omitidos_nivel,
        "formato": "siigo" if es_siigo else "plantilla",
        "errores": errores,
    }


@router.post("/limpiar")
def limpiar_catalogo(db: DB, empresa: EmpresaActiva):
    """Desactiva todas las cuentas de la empresa (soft-delete)."""
    from sqlalchemy import update as sa_update
    from db.models.catalogo import CuentaContable
    result = db.execute(
        sa_update(CuentaContable)
        .where(CuentaContable.empresa_id == empresa.id, CuentaContable.activo == True)
        .values(activo=False)
    )
    db.commit()
    return {"desactivados": result.rowcount}


@router.delete("/{codigo}", status_code=200)
def eliminar_cuenta(codigo: str, db: DB, empresa: EmpresaActiva):
    cuenta = cuentas_service.buscar_por_codigo(db, codigo, empresa_id=empresa.id)
    if not cuenta:
        raise HTTPException(404, f"Cuenta {codigo!r} no encontrada")
    db.delete(cuenta)
    db.commit()
    return {"ok": True}


@router.get("/{codigo}", response_model=CuentaOut)
def get_cuenta(codigo: str, db: DB, empresa: EmpresaActiva):
    cuenta = cuentas_service.buscar_por_codigo(db, codigo, empresa_id=empresa.id)
    if not cuenta:
        raise HTTPException(404, f"Cuenta {codigo!r} no encontrada")
    return cuenta


@router.post("/", response_model=CuentaOut, status_code=201)
def post_cuenta(body: CuentaCreate, db: DB, empresa: EmpresaActiva):
    if cuentas_service.buscar_por_codigo(db, body.codigo, empresa_id=empresa.id):
        raise HTTPException(409, f"El código {body.codigo!r} ya existe")
    cuenta = cuentas_service.crear_cuenta(
        db,
        codigo=body.codigo,
        nombre=body.nombre,
        fiscal=body.fiscal,
        empresa_id=empresa.id,
    )
    db.commit()
    return cuenta
