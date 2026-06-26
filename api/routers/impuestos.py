"""
Router: /impuestos – CRUD + carga masiva de códigos de impuesto SIIGO.
Requiere autenticación; filtra por empresa activa.
"""

from __future__ import annotations

from io import BytesIO
from typing import Annotated

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from api.dependencies import get_empresa_activa
from api.schemas import ImpuestoCreate, ImpuestoOut
from db.models.auth import Empresa
from db.session import get_db
from services import impuestos_service

router = APIRouter(prefix="/impuestos", tags=["Códigos de Impuesto"])

DB = Annotated[Session, Depends(get_db)]
EmpresaActiva = Annotated[Empresa, Depends(get_empresa_activa)]

_TIPOS_VALIDOS = {"IVA", "Impoconsumo", "Retefuente", "ReteICA", "ReteIVA"}


@router.get("/", response_model=list[ImpuestoOut])
def get_impuestos(db: DB, empresa: EmpresaActiva):
    return impuestos_service.listar_todos(db, empresa_id=empresa.id)


@router.get("/plantilla")
def descargar_plantilla(empresa: EmpresaActiva):
    """Descarga un Excel de ejemplo con el formato esperado para importación."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Impuestos"

    headers = ["codigo", "nombre", "tipo_impuesto", "tarifa", "cta_compras", "cta_ventas"]
    ejemplos = [
        ["1",  "IVA 19%",          "IVA",         19,   "24080500", ""],
        ["2",  "IVA 5%",           "IVA",          5,   "24080500", ""],
        ["3",  "Retefuente 11%",   "Retefuente",  11,   "23650500", ""],
        ["4",  "Retefuente 10%",   "Retefuente",  10,   "23650500", ""],
        ["5",  "Retefuente 6%",    "Retefuente",   6,   "23650500", ""],
        ["6",  "ReteICA 1%",       "ReteICA",      1,   "23680500", ""],
    ]

    header_fill = PatternFill("solid", fgColor="1C6B4A")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_data in ejemplos:
        ws.append(row_data)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    # Nota de tipos válidos
    ws.append([])
    nota = ws.cell(row=len(ejemplos) + 3, column=1, value=f"Tipos válidos: {', '.join(sorted(_TIPOS_VALIDOS))}")
    nota.font = Font(italic=True, color="666666")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_impuestos.xlsx"},
    )


@router.post("/cargar-excel")
async def cargar_excel(
    db: DB,
    empresa: EmpresaActiva,
    archivo: UploadFile = File(...),
):
    """Importa impuestos desde un Excel. Columna obligatoria: codigo. Hace upsert por (codigo, empresa)."""
    contenido = await archivo.read()
    try:
        df = pd.read_excel(BytesIO(contenido), dtype=str)
    except Exception as e:
        raise HTTPException(400, f"No se pudo leer el archivo: {e}")

    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

    if "codigo" not in df.columns:
        raise HTTPException(400, "El archivo debe tener una columna 'codigo'.")

    insertados = 0
    actualizados = 0
    errores: list[dict] = []

    for idx, row in df.iterrows():
        fila = int(idx) + 2  # +2 porque idx=0 y row 1 es header
        codigo = str(row.get("codigo", "")).strip()
        if not codigo:
            errores.append({"fila": fila, "error": "Código vacío"})
            continue

        def _val(col: str) -> str | None:
            v = row.get(col)
            return str(v).strip() if pd.notna(v) and str(v).strip() not in ("", "nan") else None

        def _num(col: str) -> float | None:
            v = row.get(col)
            if pd.notna(v):
                try:
                    return float(v)
                except (ValueError, TypeError):
                    pass
            return None

        try:
            _, creado = impuestos_service.upsert_impuesto(
                db,
                empresa_id=empresa.id,
                codigo=codigo,
                nombre=_val("nombre"),
                tipo_impuesto=_val("tipo_impuesto"),
                tarifa=_num("tarifa"),
                cta_compras=_val("cta_compras"),
                cta_ventas=_val("cta_ventas"),
            )
            if creado:
                insertados += 1
            else:
                actualizados += 1
        except Exception as e:
            errores.append({"fila": fila, "error": str(e)})

    db.commit()
    return {"insertados": insertados, "actualizados": actualizados, "errores": errores}


@router.get("/{codigo}", response_model=ImpuestoOut)
def get_impuesto(codigo: str, db: DB, empresa: EmpresaActiva):
    imp = impuestos_service.buscar_por_codigo(db, codigo, empresa_id=empresa.id)
    if not imp:
        raise HTTPException(404, f"Impuesto {codigo!r} no encontrado")
    return imp


@router.post("/", response_model=ImpuestoOut, status_code=201)
def post_impuesto(body: ImpuestoCreate, db: DB, empresa: EmpresaActiva):
    if impuestos_service.buscar_por_codigo(db, body.codigo, empresa_id=empresa.id):
        raise HTTPException(409, f"El código {body.codigo!r} ya existe")
    imp = impuestos_service.crear_impuesto(db, empresa_id=empresa.id, **body.model_dump())
    db.commit()
    return imp
