"""
Router: /terceros – gestión de terceros (proveedores) por empresa.
"""

from __future__ import annotations

from io import BytesIO
from typing import Annotated

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func as sqlfunc
from sqlalchemy.orm import Session

from api.dependencies import get_current_user, get_empresa_activa
from api.schemas.terceros import (
    ExportarTercerosBody,
    TerceroCatalogos,
    TerceroOut,
    TerceroUpdate,
    TipoIdentificacionOut,
    PaisOut,
    DepartamentoOut,
    CiudadOut,
    SiigoTipoPersonaOut,
    SiigoRegimenIvaOut,
    SiigoResponsabilidadFiscalOut,
)
from db.models.auth import Empresa, Usuario
from db.models.contabilidad import Proveedor
from db.models.geo import (
    TipoIdentificacion,
    Pais,
    Departamento,
    Ciudad,
    SiigoTipoPersona,
    SiigoRegimenIva,
    SiigoResponsabilidadFiscal,
)
from db.session import get_db

# ── Columnas del formato de importación masiva Siigo Nube SF_CO ───────────────

_SIIGO_HEADERS = [
    "Identificación",
    "Dígito de verificación",
    "Código Sucursal",
    "Tipo identificación",
    "Tipo",
    "Razón social",
    "Nombres del tercero",
    "Apellidos del tercero",
    "Nombre Comercial",
    "Dirección",
    "Código país",
    "Código departamento/estado",
    "Código ciudad",
    "Indicativo teléfono principal",
    "Teléfono principal",
    "Extensión teléfono principal",
    "Tipo de régimen IVA",
    "Código Responsabilidad fiscal",
    "Código Postal",
    "Nombres contacto principal",
    "Apellidos contacto principal",
    "Indicativo teléfono contacto principal",
    "Teléfono contacto principal",
    "Extensión teléfono contacto principal",
    "Correo electrónico contacto principal",
    "Clientes",
    "Estado",
]


# Mapeo de códigos DIAN/internos → etiqueta Siigo para columna "Tipo de régimen IVA"
_REGIMEN_SIIGO: dict[str, str] = {
    "0": "0 - No responsable de IVA",
    "2": "2 - Responsable de IVA",
    "RS": "0 - No responsable de IVA",   # Régimen Simplificado
    "R-99-PN": "0 - No responsable de IVA",
    "NC": "0 - No responsable de IVA",
    "RNC": "0 - No responsable de IVA",
    "RC": "2 - Responsable de IVA",       # Régimen Común
    "GC": "2 - Responsable de IVA",       # Gran Contribuyente
}

# Códigos de responsabilidad fiscal aceptados por Siigo
_RESPONSABILIDADES_VALIDAS = {"O-13", "O-15", "O-23", "O-47", "R-99-PN"}


def _prov_to_siigo_row(p: Proveedor) -> list:
    es_natural = p.tipo_persona == "natural"
    # Columna E "Tipo": solo acepta "Empresa" o "Es persona"
    tipo = "Es persona" if es_natural else "Empresa"

    if es_natural:
        # Natural: razón social vacía; nombres y apellidos van en sus propias columnas
        razon = ""
        nombres = p.nombres_tercero or ""
        apellidos = p.apellidos_tercero or ""
    else:
        razon = p.razon_social or ""
        nombres = razon  # jurídica repite razón social en nombres (campo obligatorio Siigo)
        apellidos = ""

    # Régimen IVA: mapear código interno → etiqueta Siigo
    regimen_raw = (p.tipo_regimen_iva or "").strip()
    regimen_siigo = _REGIMEN_SIIGO.get(regimen_raw, "")

    # Responsabilidad fiscal: usar campo dedicado; si tiene ZZ (inválido), buscar en
    # tipo_regimen_iva (registros viejos donde el parser almacenó R-99-PN allí)
    cod_resp = (p.codigo_responsabilidad or "").strip()
    if cod_resp in _RESPONSABILIDADES_VALIDAS:
        responsabilidad = cod_resp
    elif regimen_raw in _RESPONSABILIDADES_VALIDAS:
        responsabilidad = regimen_raw
    else:
        responsabilidad = ""

    # Correo: preferir email_contacto, caer en email principal si no hay
    correo_contacto = p.email_contacto or p.email or ""

    return [
        p.nit,
        p.digito_verificacion if p.digito_verificacion is not None else "",
        p.codigo_sucursal or "",
        p.tipo_identificacion or "",
        tipo,
        razon,
        nombres,
        apellidos,
        p.nombre_comercial or "",
        p.direccion or "",
        p.codigo_pais or "Col",
        p.codigo_departamento or "",
        p.codigo_ciudad_siigo or "",
        p.indicativo_tel or "",
        p.telefono or "",
        p.extension_tel or "",
        regimen_siigo,
        responsabilidad,
        p.codigo_postal or "",
        p.nombres_contacto or "",
        p.apellidos_contacto or "",
        p.indicativo_tel_contacto or "",
        p.telefono_contacto or "",
        p.extension_tel_contacto or "",
        correo_contacto,
        "",  # Clientes: el usuario lo completa en Siigo
        "",  # Estado: el usuario lo completa en Siigo
    ]


def _generar_excel_siigo(terceros: list[Proveedor]) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Terceros"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(_SIIGO_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.row_dimensions[1].height = 30

    for row_idx, prov in enumerate(terceros, start=2):
        for col_idx, value in enumerate(_prov_to_siigo_row(prov), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

router = APIRouter(prefix="/terceros", tags=["Terceros"])

DB = Annotated[Session, Depends(get_db)]
EmpresaActiva = Annotated[Empresa, Depends(get_empresa_activa)]
CurrentUser = Annotated[Usuario, Depends(get_current_user)]


# ── Catálogos de referencia ───────────────────────────────────────────────────

@router.get("/catalogos", response_model=TerceroCatalogos)
def get_catalogos(db: DB, _: CurrentUser):
    """Retorna los catálogos necesarios para los formularios del módulo."""
    tipos = db.scalars(
        select(TipoIdentificacion)
        .where(TipoIdentificacion.activo == True)
        .order_by(TipoIdentificacion.codigo)
    ).all()
    paises = db.scalars(
        select(Pais).order_by(Pais.nombre)
    ).all()
    tipos_persona = db.scalars(
        select(SiigoTipoPersona).order_by(SiigoTipoPersona.id)
    ).all()
    regimenes_iva = db.scalars(
        select(SiigoRegimenIva).order_by(SiigoRegimenIva.codigo)
    ).all()
    responsabilidades = db.scalars(
        select(SiigoResponsabilidadFiscal).order_by(SiigoResponsabilidadFiscal.id)
    ).all()
    return TerceroCatalogos(
        tipos_identificacion=[TipoIdentificacionOut.model_validate(t) for t in tipos],
        paises=[PaisOut.model_validate(p) for p in paises],
        tipos_persona=[SiigoTipoPersonaOut.model_validate(tp) for tp in tipos_persona],
        regimenes_iva=[SiigoRegimenIvaOut.model_validate(r) for r in regimenes_iva],
        responsabilidades_fiscales=[SiigoResponsabilidadFiscalOut.model_validate(r) for r in responsabilidades],
    )


@router.get("/departamentos", response_model=list[DepartamentoOut])
def get_departamentos(
    db: DB,
    _: CurrentUser,
    pais_codigo: str = Query(default="Col"),
):
    """Lista departamentos/estados de un país."""
    rows = db.scalars(
        select(Departamento)
        .where(Departamento.pais_codigo == pais_codigo)
        .order_by(Departamento.nombre)
    ).all()
    return [DepartamentoOut.model_validate(r) for r in rows]


@router.get("/ciudades", response_model=list[CiudadOut])
def buscar_ciudades(
    db: DB,
    _: CurrentUser,
    q: str = Query(default="", min_length=0),
    departamento_codigo: str | None = Query(default=None),
    pais_codigo: str = Query(default="Col"),
    limit: int = Query(default=20, le=100),
):
    """Busca ciudades por nombre (útil para autocomplete)."""
    stmt = select(Ciudad).where(Ciudad.pais_codigo == pais_codigo)
    if departamento_codigo:
        stmt = stmt.where(Ciudad.departamento_codigo == departamento_codigo)
    if q:
        stmt = stmt.where(
            sqlfunc.lower(Ciudad.nombre).contains(q.lower())
        )
    stmt = stmt.order_by(Ciudad.nombre).limit(limit)
    rows = db.scalars(stmt).all()
    return [CiudadOut.model_validate(r) for r in rows]


# ── CRUD de terceros ──────────────────────────────────────────────────────────

@router.get("", response_model=list[TerceroOut])
def listar_terceros(
    db: DB,
    empresa: EmpresaActiva,
    _: CurrentUser,
    q: str | None = Query(default=None),
    tipo_persona: str | None = Query(default=None),
    solo_activos: bool = Query(default=True),
    limit: int = Query(default=100, le=500),
    offset: int = Query(default=0),
):
    """Lista los terceros registrados para la empresa activa."""
    stmt = (
        select(Proveedor)
        .where(Proveedor.empresa_id == empresa.id)
        .order_by(Proveedor.razon_social)
    )
    if solo_activos:
        stmt = stmt.where(Proveedor.activo == True)
    if tipo_persona:
        stmt = stmt.where(Proveedor.tipo_persona == tipo_persona)
    if q:
        q_low = q.lower()
        from sqlalchemy import or_
        stmt = stmt.where(
            or_(
                sqlfunc.lower(Proveedor.nit).contains(q_low),
                sqlfunc.lower(Proveedor.razon_social).contains(q_low),
            )
        )
    stmt = stmt.offset(offset).limit(limit)
    rows = db.scalars(stmt).all()
    return [TerceroOut.model_validate(r) for r in rows]


@router.get("/{tercero_id}", response_model=TerceroOut)
def get_tercero(
    tercero_id: int,
    db: DB,
    empresa: EmpresaActiva,
    _: CurrentUser,
):
    """Obtiene el detalle completo de un tercero."""
    prov = db.scalar(
        select(Proveedor).where(
            Proveedor.id == tercero_id,
            Proveedor.empresa_id == empresa.id,
        )
    )
    if prov is None:
        raise HTTPException(404, "Tercero no encontrado")
    return TerceroOut.model_validate(prov)


@router.delete("/{tercero_id}", status_code=204)
def eliminar_tercero(
    tercero_id: int,
    db: DB,
    empresa: EmpresaActiva,
    _: CurrentUser,
):
    """Desactiva un tercero (lo oculta de listas y exportaciones)."""
    prov = db.scalar(
        select(Proveedor).where(
            Proveedor.id == tercero_id,
            Proveedor.empresa_id == empresa.id,
        )
    )
    if prov is None:
        raise HTTPException(404, "Tercero no encontrado")
    prov.activo = False
    db.commit()


@router.put("/{tercero_id}", response_model=TerceroOut)
def actualizar_tercero(
    tercero_id: int,
    body: TerceroUpdate,
    db: DB,
    empresa: EmpresaActiva,
    _: CurrentUser,
):
    """Actualiza los datos de un tercero (edición manual desde la UI)."""
    prov = db.scalar(
        select(Proveedor).where(
            Proveedor.id == tercero_id,
            Proveedor.empresa_id == empresa.id,
        )
    )
    if prov is None:
        raise HTTPException(404, "Tercero no encontrado")

    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(prov, field, value)

    # Recalcular DV si cambió el tipo_identificacion
    if body.tipo_identificacion is not None:
        from core.dv import calcular_dv
        prov.digito_verificacion = calcular_dv(prov.nit, prov.tipo_identificacion)

    db.commit()
    db.refresh(prov)
    return TerceroOut.model_validate(prov)


# ── Exportación formato Siigo ─────────────────────────────────────────────────

@router.post("/exportar")
def exportar_terceros(
    body: ExportarTercerosBody,
    db: DB,
    empresa: EmpresaActiva,
    _: CurrentUser,
):
    """
    Genera un Excel en el formato de importación masiva de Siigo Nube SF_CO.
    - ids vacío → exporta todos los terceros activos de la empresa.
    - ids con valores → exporta solo los terceros con esos IDs.
    """
    stmt = (
        select(Proveedor)
        .where(Proveedor.empresa_id == empresa.id, Proveedor.activo == True)
        .order_by(Proveedor.razon_social)
    )
    if body.ids:
        stmt = stmt.where(Proveedor.id.in_(body.ids))

    terceros = db.scalars(stmt).all()

    buf = _generar_excel_siigo(list(terceros))
    filename = "terceros_siigo.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Estadísticas rápidas ──────────────────────────────────────────────────────

@router.get("/stats/resumen")
def stats_terceros(db: DB, empresa: EmpresaActiva, _: CurrentUser):
    """Resumen estadístico del módulo de terceros."""
    total = db.scalar(
        select(sqlfunc.count()).where(
            Proveedor.empresa_id == empresa.id,
            Proveedor.activo == True,
        )
    ) or 0
    juridicas = db.scalar(
        select(sqlfunc.count()).where(
            Proveedor.empresa_id == empresa.id,
            Proveedor.activo == True,
            Proveedor.tipo_persona == "juridica",
        )
    ) or 0
    naturales = db.scalar(
        select(sqlfunc.count()).where(
            Proveedor.empresa_id == empresa.id,
            Proveedor.activo == True,
            Proveedor.tipo_persona == "natural",
        )
    ) or 0
    completos = db.scalar(
        select(sqlfunc.count()).where(
            Proveedor.empresa_id == empresa.id,
            Proveedor.activo == True,
            Proveedor.tipo_identificacion.is_not(None),
            Proveedor.codigo_departamento.is_not(None),
        )
    ) or 0
    return {
        "total": total,
        "juridicas": juridicas,
        "naturales": naturales,
        "completos": completos,
        "pct_completos": round(completos / total * 100, 1) if total else 0,
    }
